'''
File Name      : activity_data_daily.py
Author Name    : Lakshmi Damodara
Date           : 01/28/2018
Version        : 1.0
Description    :
This program reads the individual sheets in mechanical tracker file and gets the values for the daily activity Data table
The hard-coded value is the directory and the file of the excel_act_tble_config.ini file.
The cell positions are given in the config file.

The program reads the various attributes of the cell for activity columns and writes to output file
using the program excel_writing.py

Functions:
converDate, convertDate1

Files need to run this program:
1. excel_file_config.ini

Program dependencies:
1. excel_file_config_reader.py
2. excel_writing.py

Log File
1. log_file.txt : has been set at the DEBUG level to log all activities for the run time.

Output File
1. Based on the number of activities, respective files are written in the output directory
2. The output directory can be configured on excel_file_config.ini file

'''

import openpyxl
import configparser
import datetime
import logging


import excel_writing as ewWriter
import Database_Insert as dbi

### -- Start of Functions --------
# function to convert dates to string in mmddyyyy format
def convertDate(dtt):
    return datetime.datetime.date(dtt) # returns just the date in mm-dd-yyyy format

def convertDate1(dtt):
    return dtt.strftime('%m%d%Y') # returns date in string in mmddyyyy format

# returns incremented value of a variable
def incrementfnc(tval):
    tval = tval + 1
    return tval

def remove_items_list(listVar,removeListVar):
    # remove the unnecessary values from the result data
    for i in sorted(removeListVar, reverse=True):
        del listVar[i]
    return listVar

#---------------------------------------------------------------------------
# This function contains the business logic to read the daily activity sheet
#----------------------------------------------------------------------------
def getSheetResult(wbe,active_sheet_value):
    sName = efcr.shName(active_sheet_value)  # get the sheet name
    Asheet = wbe[sName]
   # get the max count of rows and cols
    m_row = Asheet.max_row
    m_col = Asheet.max_column
    m_row = m_row + 1

    # -- This section is to store the data of the active sheet in to List of Lists
    # -- result_data is the list of list containing : rows and columns of the active sheet
    # -- The reading of row starts at Row# 7 and 6 columns are read starting from column # 2
    # -- As it is being read, the dates are converted to MMDDYYYY format and
    # -- the result is stored in result_data as list of lists

    # initialize list
    result_data = []
    for curr_row in range(7, m_row, 1):
        if not Asheet.row_dimensions[curr_row].hidden == True:  # dont read if the row is hidden
            row_data = []
            row_data.append(sName) # inserting the activity name
            for curr_col in range(2, 8, 1): # read each col. from the sheet starting from col number 2 upto col 8
                data = Asheet.cell(row=curr_row, column=curr_col)
                if isinstance(data.value, datetime.datetime): # getting the date value and converting to mmddyyyy format value
                    row_data.append(convertDate(data.value).strftime('%m%d%Y')) # inserting the value to row_data
                else:
                    row_data.append(data.value) # inserting the rest of the values

        result_data.append(row_data) # inserting the row_data into result_data list

    # -- This section is to create a list - popping_Var which contains the row index of result_data
    # -- that needed to be removed as it contains non-date values
    # -- We are reading the col.#2 and checking if the string value is greater than 10
    # -- If yes, then the index value is stored in popping_Var

    popping_Var = []
    ## Now accessing the list of list result_data : result_data = [][]
    for i in range(0, len(result_data), 1):
        for j in range(1, 7, 1):  # access the list from index of 1 in result_data as the index[0] is the project name
            if type(result_data[i][j]) == str and len(result_data[i][j]) > 10:
                popping_Var.append(i)
                break

    # Call the function to remove the list of values in result_data which are referenced in popping_Var
    result_data = remove_items_list(result_data, popping_Var)

### -------------------------------------
### Remove the list of result Data which has None or null values
### If the planned to date column has null values, those rows in the result_data list are removed.
### --------------------------------------
    popping_Var_None = []
    ## Now accessing the list of list result_data with None Value
    for i in range(0, len(result_data), 1):
        if result_data[i][5] == None: # checking if the list index[5] in result_data is None
            popping_Var_None.append(i) # store the index value in popping_Var_None list

    # Call the function to remove the list containing None or null values as referenced in popping_Var_None
    result_data=remove_items_list(result_data,popping_Var_None)

    # return the final list to the calling function
    logging.info(result_data)
    del popping_Var
    del popping_Var_None
    return result_data


### ---------End of Functions -----

# import excel_file_config_readyer.py to get all its functions
import excel_file_config_reader as efcr

# get the filename and directory for logfile writing
Log_FileName = efcr.logfileDirectory() + efcr.logfileName()  # directory + filename
logging.basicConfig(filename=Log_FileName,level=logging.DEBUG,)

logging.info('##---Program: activities_data_daily.py..........................')
logging.info(datetime.datetime.today())
logging.info('##-------------------- ---------------------------------........')

# get the filename and directory : Excel fileName and directory for reading values
L_FileName = efcr.fileDirectory() + efcr.fileName() # directory + filename
logging.info('activities_data_daily.py : opening excel file name - %s'%L_FileName)
print(L_FileName)
# passing the file name and creating an instance of the workbook with actual values and ignoring the formulas
wb = openpyxl.load_workbook(L_FileName,data_only='True')

# Fist get the active range sheets from excel_file_config.ini using excel_file_config_reader.py
asheets = []
result_data_sheet = []

#----------------------------------------------------------------------------------------------
# -- This section is to collect all the data from various activity sheet
# -- Call the excel writing.py program to write the file into a csv format
# -- The information about sheetname, total sheets etc are derived from excel_file_config.ini
# -- The output file is determined by the name of the sheet.csv
#----------------------------------------------------------------------------------------------

asheets = efcr.getActivitySheets() # get the list of activity sheets from excel_file_config.ini
len_asheets = len((asheets))
for i in range(0,len_asheets,2):
    sheet_val = asheets[i] # getting the active worksheet number
    result_data_sheet = getSheetResult(wb,sheet_val) # calling function getSheetResult()
    loc_fname = efcr.outputDirectory() + efcr.outputfileName() # getting the output directory and filename
    # calling the excel_writing.py to write the data to the file
    ewWriter.write_activity_daily_data_CSV(loc_fname, result_data_sheet)
    dbi.executeSQL_Activities_Daily(result_data_sheet)

# close all the connections
wb.close()
# close or delete all the open instances, Lists, and connections
# clears all the variables from memory

del result_data_sheet
del asheets
del efcr
del ewWriter
del dbi

#---- End of Program ------